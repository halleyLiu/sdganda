#!/usr/bin/env python3
"""
Fetch SDGanda unit data plus 14-day win-rate data and write daily JSON/XLSX files.

Default output:
  ./YYYYMMDD/
    units.json
    win_rates.json
    combined_units_win_rate.json
    sdganda_units_YYYYMMDD.xlsx
"""

from __future__ import annotations

import argparse
import concurrent.futures
import datetime as dt
import json
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


API_BASE_URL = "https://dbapi.sdganda.com/api"
RANK_LABELS = ["C", "C", "B", "A", "S"]
RANK_TYPE_SUFFIX = {3: "S", 4: "R", 5: "U"}


def fetch_json(url: str, retries: int = 3, timeout: int = 30) -> dict[str, Any]:
    last_error: Exception | None = None
    headers = {"User-Agent": "sdganda-units-export/1.0"}

    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(min(2 * attempt, 6))

    raise RuntimeError(f"Failed to fetch {url}: {last_error}")


def rank_display(rank: Any, rank_type: Any) -> str:
    try:
        rank_value = int(rank)
    except (TypeError, ValueError):
        return ""

    base = RANK_LABELS[rank_value] if 0 <= rank_value < len(RANK_LABELS) else str(rank)

    try:
        type_value = int(rank_type)
    except (TypeError, ValueError):
        return base

    return f"{base}{RANK_TYPE_SUFFIX.get(type_value, '')}"


def fetch_units(api_base_url: str, limit: int) -> list[dict[str, Any]]:
    url = f"{api_base_url.rstrip('/')}/units?{urllib.parse.urlencode({'limit': limit})}"
    payload = fetch_json(url)
    if not payload.get("success") or not isinstance(payload.get("data"), list):
        raise RuntimeError(f"Unexpected units response: {payload}")
    return payload["data"]


def fetch_win_rate(api_base_url: str, unit_id: int, days: int) -> dict[str, Any]:
    url = f"{api_base_url.rstrip('/')}/sdgore/win-rate/{unit_id}/{days}"
    payload = fetch_json(url)
    if payload.get("success") and payload.get("data", {}).get("code") == 0:
        data = payload.get("data", {}).get("data")
        if isinstance(data, dict):
            return data
    return {
        "unit_id": unit_id,
        "win_times": None,
        "lose_times": None,
        "draw_times": None,
        "win_rate": None,
        "lose_rate": None,
        "error": payload,
    }


def is_scalar(value: Any) -> bool:
    return value is None or isinstance(value, (str, int, float, bool))


def combine_unit(unit: dict[str, Any], win_rate: dict[str, Any] | None) -> dict[str, Any]:
    row: dict[str, Any] = {
        "ID": unit.get("ID"),
        "Name": unit.get("Name"),
        "品阶": rank_display(unit.get("Rank"), unit.get("RankType")),
        "win_times": None,
        "lose_times": None,
        "draw_times": None,
        "win_rate": None,
        "lose_rate": None,
    }

    for key, value in unit.items():
        if is_scalar(value) and key not in row:
            row[key] = value

    if win_rate:
        for key in ("win_times", "lose_times", "draw_times", "win_rate", "lose_rate"):
            row[key] = win_rate.get(key)
        if win_rate.get("error") is not None:
            row["win_rate_error"] = json.dumps(win_rate["error"], ensure_ascii=False)

    rating = unit.get("rating")
    if isinstance(rating, dict):
        row["averageRating"] = rating.get("averageRating")
        row["totalRatings"] = rating.get("totalRatings")

    return row


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_excel(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "units_win_rate"

    preferred = [
        "ID",
        "Name",
        "EngName",
        "品阶",
        "Rank",
        "RankType",
        "Type",
        "win_times",
        "lose_times",
        "draw_times",
        "win_rate",
        "lose_rate",
        "averageRating",
        "totalRatings",
    ]
    all_keys = []
    seen = set()
    for key in preferred:
        if any(key in row for row in rows):
            all_keys.append(key)
            seen.add(key)
    for row in rows:
        for key in row:
            if key not in seen:
                all_keys.append(key)
                seen.add(key)

    ws.append(all_keys)
    for row in rows:
        ws.append([row.get(key) for key in all_keys])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    percent_columns = {"win_rate", "lose_rate"}
    for idx, header in enumerate(all_keys, start=1):
        letter = get_column_letter(idx)
        width = min(max(len(str(header)) + 2, 10), 30)
        for cell in ws[2 : min(ws.max_row, 80)]:
            value = cell[idx - 1].value
            if value is not None:
                width = min(max(width, len(str(value)) + 2), 40)
        ws.column_dimensions[letter].width = width
        if header in percent_columns:
            for cell in ws[letter][1:]:
                cell.number_format = "0.00%"

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export SDGanda units and win-rate data.")
    parser.add_argument("--api-base-url", default=API_BASE_URL)
    parser.add_argument("--days", type=int, default=14, help="Win-rate day window.")
    parser.add_argument("--limit", type=int, default=1000, help="Units API page size.")
    parser.add_argument("--workers", type=int, default=8, help="Concurrent win-rate requests.")
    parser.add_argument("--output-root", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--date", default=dt.datetime.now().strftime("%Y%m%d"))
    parser.add_argument("--sample", type=int, default=0, help="Fetch only first N units for testing.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    out_dir = args.output_root / args.date
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Fetching units from {args.api_base_url}/units ...")
    units = fetch_units(args.api_base_url, args.limit)
    if args.sample:
        units = units[: args.sample]
        print(f"Sample mode: only fetching win-rate for {len(units)} units.")
    else:
        print(f"Fetched {len(units)} units.")

    win_rates: dict[str, dict[str, Any]] = {}
    print(f"Fetching {args.days}-day win-rate data with {args.workers} workers ...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as executor:
        future_to_id = {
            executor.submit(fetch_win_rate, args.api_base_url, int(unit["ID"]), args.days): int(unit["ID"])
            for unit in units
            if unit.get("ID") is not None
        }
        for index, future in enumerate(concurrent.futures.as_completed(future_to_id), start=1):
            unit_id = future_to_id[future]
            try:
                win_rates[str(unit_id)] = future.result()
            except Exception as exc:
                win_rates[str(unit_id)] = {
                    "unit_id": unit_id,
                    "win_times": None,
                    "lose_times": None,
                    "draw_times": None,
                    "win_rate": None,
                    "lose_rate": None,
                    "error": str(exc),
                }
            if index % 50 == 0 or index == len(future_to_id):
                print(f"  win-rate progress: {index}/{len(future_to_id)}")

    rows = [combine_unit(unit, win_rates.get(str(unit.get("ID")))) for unit in units]

    units_path = out_dir / "units.json"
    win_rates_path = out_dir / "win_rates.json"
    combined_path = out_dir / "combined_units_win_rate.json"
    excel_path = out_dir / f"sdganda_units_{args.date}.xlsx"

    write_json(units_path, units)
    write_json(win_rates_path, win_rates)
    write_json(combined_path, rows)
    write_excel(excel_path, rows)

    print(f"Wrote: {units_path}")
    print(f"Wrote: {win_rates_path}")
    print(f"Wrote: {combined_path}")
    print(f"Wrote: {excel_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
